import logging
import os
import shutil
import subprocess
import tempfile
from datetime import datetime, timedelta
from num2words import num2words
from docxtpl import DocxTemplate
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")

def garantir_pastas():
    os.makedirs(TEMPLATES_DIR, exist_ok=True)

def valor_por_extenso(valor_str):
    if not valor_str:
        return ""
    try:
        limpo = valor_str.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        valor_float = float(limpo)
        extenso = num2words(valor_float, lang='pt_BR', to='currency')
        return f"({extenso.capitalize()})"
    except Exception as e:
        print(f"Erro ao converter extenso: {e}")
        return ""

def criar_contexto_cliente(cliente: dict) -> dict:
    hoje = datetime.now()
    data_60_dias = hoje + timedelta(days=60)
    
    valor_obra = cliente.get("valor_da_obra") or "0,00"
    valor_devolucao = cliente.get("valor_de_devolucao") or "0,00"
    
    return {
        "id": cliente.get("id", ""),
        "nome": cliente.get("nome", ""),
        "cpf_cnpj": cliente.get("cpf_cnpj", ""),
        "endereco": cliente.get("endereco", ""),
        "cidade": cliente.get("cidade", ""),
        "cep": cliente.get("cep", ""),
        "nota_ps": cliente.get("nota_ps", ""),
        "valor_de_devolucao": valor_devolucao,
        "valor_devolucao_extenso": valor_por_extenso(valor_devolucao),
        "valor_da_obra": valor_obra,
        "valor_extenso": valor_por_extenso(valor_obra),
        "data": hoje.strftime("%d/%m/%Y"),
        "data_fim": data_60_dias.strftime("%d/%m/%Y")
    }

def convert_docx_to_pdf(docx_path: str, out_dir: str) -> str:
    """Converte arquivo DOCX para PDF usando LibreOffice Headless ou Word COM (Windows).
    
    Lança RuntimeError com mensagem descritiva se a conversão falhar.
    """
    pdf_name = os.path.splitext(os.path.basename(docx_path))[0] + ".pdf"
    pdf_path_esperado = os.path.join(out_dir, pdf_name)
    erros = []

    # 1. Procura por libreoffice no PATH (Linux / Mac / Windows configurado)
    libreoffice_bin = shutil.which("libreoffice") or shutil.which("soffice")

    # 2. Tenta caminhos padrão no Windows se não estiver no PATH
    if not libreoffice_bin and os.name == 'nt':
        windows_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for path in windows_paths:
            if os.path.exists(path):
                libreoffice_bin = path
                break

    if libreoffice_bin:
        logger.info("Convertendo DOCX→PDF via LibreOffice: %s", libreoffice_bin)
        try:
            result = subprocess.run(
                [
                    libreoffice_bin,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    out_dir,
                    docx_path,
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=60,
            )
            stdout = result.stdout.decode("utf-8", errors="ignore")
            stderr = result.stderr.decode("utf-8", errors="ignore")
            logger.debug("LibreOffice stdout: %s", stdout)
            if result.returncode != 0:
                logger.warning("LibreOffice retornou código %d. stderr: %s", result.returncode, stderr)
                erros.append(f"LibreOffice (código {result.returncode}): {stderr.strip() or stdout.strip()}")
            # Verifica se o PDF foi gerado mesmo com returncode != 0 (comportamento comum do LO)
            if os.path.exists(pdf_path_esperado):
                logger.info("PDF gerado com sucesso via LibreOffice: %s", pdf_path_esperado)
                return pdf_path_esperado
            else:
                msg = f"LibreOffice executou mas o PDF não foi criado em '{pdf_path_esperado}'. {stderr.strip()}"
                logger.error(msg)
                erros.append(msg)
        except subprocess.TimeoutExpired:
            msg = "LibreOffice excedeu o tempo limite de 60 segundos."
            logger.error(msg)
            erros.append(msg)
        except Exception as exc:
            msg = f"Exceção ao executar LibreOffice: {exc}"
            logger.exception(msg)
            erros.append(msg)
    else:
        msg = "LibreOffice não encontrado no PATH nem nos caminhos padrão do Windows."
        logger.warning(msg)
        erros.append(msg)

    # 3. Fallback: Automação COM do Word (Apenas Windows)
    if os.name == 'nt':
        logger.info("Tentando fallback via Word COM (pywin32)...")
        try:
            import pythoncom
            import win32com.client

            pythoncom.CoInitialize()
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            try:
                abs_docx = os.path.abspath(docx_path)
                pdf_path = os.path.join(out_dir, pdf_name)
                doc = word.Documents.Open(abs_docx)
                doc.SaveAs(pdf_path, FileFormat=17)  # 17 = wdFormatPDF
                doc.Close()
                if os.path.exists(pdf_path):
                    logger.info("PDF gerado com sucesso via Word COM: %s", pdf_path)
                    return pdf_path
                else:
                    msg = "Word COM executou mas o PDF não foi criado."
                    logger.error(msg)
                    erros.append(msg)
            finally:
                word.Quit()
                pythoncom.CoUninitialize()
        except ImportError:
            msg = "pywin32 não instalado. Instale com: pip install pywin32"
            logger.warning(msg)
            erros.append(msg)
        except Exception as exc:
            msg = f"Erro no fallback Word COM: {exc}"
            logger.exception(msg)
            erros.append(msg)

    # Nenhum método funcionou — lança exceção com todos os detalhes
    detalhe = " | ".join(erros) if erros else "Nenhum conversor disponível (LibreOffice ou pywin32)."
    raise RuntimeError(f"Falha ao converter DOCX para PDF. Detalhes: {detalhe}")

def preencher_word(cliente: dict, template_name: str, out_dir: str) -> str:
    garantir_pastas()
    contexto = criar_contexto_cliente(cliente)
    template_path = os.path.join(TEMPLATES_DIR, f"{template_name}.docx")
    
    if not os.path.exists(template_path):
        return None
        
    doc = DocxTemplate(template_path)
    doc.render(contexto)
    
    nome_cliente = "".join([c for c in str(contexto['nome']) if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
    nome_template = "".join([c for c in str(template_name) if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
    nome_arq = f"{nome_template} - {nome_cliente}.docx"
    caminho = os.path.join(out_dir, nome_arq)
    doc.save(caminho)
    return caminho

def preencher_excel(cliente: dict, template_name: str, out_dir: str) -> str:
    garantir_pastas()
    contexto = criar_contexto_cliente(cliente)
    template_path = os.path.join(TEMPLATES_DIR, f"{template_name}.xlsx")
    
    if not os.path.exists(template_path):
        return None
        
    wb = load_workbook(template_path)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                valor_original = cell.value
                for chave, valor in contexto.items():
                    placeholder = f"{{{{{chave}}}}}"
                    if placeholder in valor_original:
                        cell.value = valor_original.replace(placeholder, str(valor))
                        valor_original = cell.value
                        
    nome_cliente = "".join([c for c in str(contexto['nome']) if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
    nome_template = "".join([c for c in str(template_name) if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
    nome_arq = f"{nome_template} - {nome_cliente}.xlsx"
    caminho = os.path.join(out_dir, nome_arq)
    wb.save(caminho)
    return caminho
