"""Cadastros de apoio do módulo Controle de O.S.: Obras, Equipes e Produtos.

Segue o mesmo padrão dos demais routers do projeto: Supabase via PostgREST,
validação com Pydantic e permissão de módulo ("os").
"""

import io
import logging
import re
import unicodedata

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field, field_validator

from auth import require_permisao, require_qualquer_permisao
from supabase_client import get_supabase
from utils.tipos_os import TIPOS_OS

# Alias legado usado nas validações de contrato deste módulo.
TIPOS_SERVICO = TIPOS_OS

# Catálogo de produtos é leitura necessária ao usuário de campo (lançamento de
# serviços na O.S); as demais operações de cadastro seguem restritas ao gestor.
router = APIRouter(dependencies=[Depends(require_qualquer_permisao(["os", "os_campo"]))])

GESTOR_ONLY = [Depends(require_permisao("os"))]

logger = logging.getLogger(__name__)


def _termo_busca_seguro(termo: str | None) -> str:
    """Remove caracteres da gramática do PostgREST (or_/ilike) que, digitados
    pelo usuário, derrubariam a busca com 500: `( ) , . % : ; "` e colchetes."""
    if not termo:
        return ""
    return re.sub(r"[(),.%:;\"\\\[\]]", "", termo).strip()

# ---------------------------------------------------------------------------
# Schemas Pydantic
# ---------------------------------------------------------------------------


class ObraCreate(BaseModel):
    # Cliente do cadastro OU Cliente Celesc (obra de terceiro): um dos dois
    # deve ser informado, mas o cliente_id é opcional no banco.
    cliente_id: int | None = Field(None, description="ID do cliente dono da obra (cadastro de clientes)")
    cliente_celesc: str | None = Field(None, max_length=255, description="Nome/contrato quando a obra é da Celesc (sem cadastro de cliente)")
    nome: str = Field(..., min_length=2, description="Nome/identificação da obra")
    endereco: str | None = None
    cidade: str | None = None

    @field_validator("cliente_celesc")
    @classmethod
    def _validar_celesc(cls, v):
        if isinstance(v, str):
            v = v.strip() or None
        return v


class ClienteMinResponse(BaseModel):
    nome: str


class ObraResponse(ObraCreate):
    id: int
    ativo: bool = True
    created_at: str | None = None
    clientes: ClienteMinResponse | None = None


class EquipeCreate(BaseModel):
    nome: str = Field(..., min_length=2)
    numero: str | None = Field(None, max_length=20, description="Número impresso no modelo de O.S (ex.: 12204)")
    descricao: str | None = None
    # IDs dos funcionários que compõem a equipe
    membro_ids: list[int] = Field(default_factory=list)
    # ID do líder: precisa pertencer à lista de membros (validado no endpoint)
    lider_id: int | None = None


class EquipeResponse(BaseModel):
    id: int
    nome: str
    numero: str | None
    descricao: str | None
    ativa: bool
    membros: list[dict]


class ProdutoCreate(BaseModel):
    nome: str = Field(..., min_length=2, description="Nome/descrição do serviço")
    codigo: str | None = Field(None, description="SKU/código de barras p/ bipagem (USC normal)")
    codigo_especial: str | None = Field(None, description="Código do serviço quando aplicado como USC especial")
    unidade: str = Field("UN", max_length=20)
    # Qtd USC (quantidade de unidades de serviço de construção) — a coluna
    # física continua `preco_unitario` (reaproveitada); o valor é exibido como USC.
    preco_unitario: float = Field(0, ge=0, description="Qtd USC")
    qtd_usc_especial: float = Field(0, ge=0, description="Qtd USC especial (adicional)")
    # Contrato (tipo de O.S) dono do serviço: construcao, manutencao ou linha_viva.
    # Obrigatório ao criar/editar; NULL só existe em registros legados.
    tipo: str = Field(..., description="Contrato do serviço: 'construcao', 'manutencao' ou 'linha_viva'")

    @field_validator("codigo", "codigo_especial")
    @classmethod
    def _normalizar_codigo(cls, v):
        if isinstance(v, str):
            v = v.strip() or None
        return v

    @field_validator("nome", "unidade")
    @classmethod
    def _normalizar_texto(cls, v):
        if isinstance(v, str):
            v = v.strip()
        if v is None or v == "":
            raise ValueError("Não pode ficar em branco.")
        return v


class ProdutoResponse(BaseModel):
    id: int
    ativo: bool = True  # default do banco
    nome: str
    codigo: str | None = None
    codigo_especial: str | None = None
    unidade: str = "UN"
    preco_unitario: float = 0  # Qtd USC
    qtd_usc_especial: float = 0  # default do banco
    # NULL = legado (disponível em todos os contratos)
    tipo: str | None = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _obter_ou_404(db, tabela: str, registro_id: int, rotulo: str) -> dict:
    resp = db.table(tabela).select("*").eq("id", registro_id).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail=f"{rotulo} não encontrado(a).")
    return resp.data[0]


def _validar_cliente_obra(db, obra: ObraCreate) -> dict:
    """Normaliza o cliente da obra (cadastro OU Celesc) antes de gravar."""
    if obra.cliente_id and obra.cliente_celesc:
        # Nunca os dois: o cadastro tem precedência (evita divergência).
        raise HTTPException(
            status_code=400,
            detail="Informe o cliente do cadastro OU o cliente Celesc, não ambos.",
        )
    if not obra.cliente_id and not obra.cliente_celesc:
        raise HTTPException(
            status_code=400,
            detail="Informe o cliente (cadastro de clientes ou Cliente Celesc).",
        )
    dados = obra.model_dump()
    if obra.cliente_id:
        _obter_ou_404(db, "clientes", obra.cliente_id, "Cliente")
        dados["cliente_celesc"] = None
    return dados


def _validar_membros_existem(db, membro_ids: list[int]) -> None:
    """Confirma que todos os funcionários existem ANTES de tocar a equipe.

    Sem esta checagem prévia, uma equipe poderia ser criada/atualizada e só
    depois a composição falhar por FK — deixando equipe órfã ou estado parcial.
    """
    ids = list(dict.fromkeys(membro_ids or []))
    if not ids:
        return
    encontrados = db.table("funcionarios").select("id").in_("id", ids).execute().data or []
    existentes = {f["id"] for f in encontrados}
    ausentes = [fid for fid in ids if fid not in existentes]
    if ausentes:
        raise HTTPException(
            status_code=400,
            detail=f"Funcionário(s) não encontrado(s): {', '.join(map(str, ausentes))}.",
        )


def _gravar_membros(db, equipe_id: int, membro_ids: list[int], lider_id: int | None):
    """Regrava a composição da equipe de forma consistente (sem estado parcial).

    - Valida a EXISTÊNCIA dos funcionários antes de tocar a equipe (evita
      metade do time gravado e o resto recusado por FK);
    - Aplica a mudança como DIFERENÇA: membros NOVOS entram primeiro e só
      depois os REMOVIDOS saem — se a gravação dos novos falhar, a equipe
      permanece com a composição anterior (nada é perdido);
    - Mantidos/adicionados recebem a flag de líder conforme `lider_id`.
    """
    membro_ids = list(dict.fromkeys(membro_ids))  # dedup preservando ordem
    if lider_id is not None and lider_id not in membro_ids:
        raise HTTPException(status_code=400, detail="O líder deve ser um membro da equipe.")

    if membro_ids:
        encontrados = db.table("funcionarios").select("id").in_("id", membro_ids).execute().data or []
        existentes = {f["id"] for f in encontrados}
        ausentes = [fid for fid in membro_ids if fid not in existentes]
        if ausentes:
            raise HTTPException(
                status_code=400,
                detail=f"Funcionário(s) não encontrado(s): {', '.join(map(str, ausentes))}.",
            )

    atuais = db.table("equipe_membros").select("id, funcionario_id").eq("equipe_id", equipe_id).execute().data or []
    atuais_por_func = {m["funcionario_id"]: m["id"] for m in atuais}

    para_adicionar = [fid for fid in membro_ids if fid not in atuais_por_func]
    para_remover = [fid for fid in atuais_por_func if fid not in membro_ids]

    # 1) Novos membros (com o time antigo intacto se isto falhar).
    if para_adicionar:
        linhas = [
            {"equipe_id": equipe_id, "funcionario_id": fid, "lider": fid == lider_id}
            for fid in para_adicionar
        ]
        resp = db.table("equipe_membros").insert(linhas).execute()
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao salvar membros da equipe.")

    # 2) Ajusta a flag de líder nos membros mantidos (ex.: remanejamento).
    for fid in membro_ids:
        if fid in atuais_por_func:
            db.table("equipe_membros").update({"lider": fid == lider_id}).eq("id", atuais_por_func[fid]).execute()

    # 3) Só então remove quem saiu.
    for fid in para_remover:
        db.table("equipe_membros").delete().eq("id", atuais_por_func[fid]).execute()


def _membros_da_equipe(db, equipe_id: int) -> list[dict]:
    """Retorna os membros da equipe com nome resolvido via join manual."""
    resp = (
        db.table("equipe_membros")
        .select("id, funcionario_id, lider, funcionarios(nome, cpf)")
        .eq("equipe_id", equipe_id)
        .execute()
    )
    return [
        {
            "id": m["id"],
            "funcionario_id": m["funcionario_id"],
            "nome": (m.get("funcionarios") or {}).get("nome"),
            "lider": m.get("lider", False),
        }
        for m in resp.data
    ]


# ---------------------------------------------------------------------------
# Obras
# ---------------------------------------------------------------------------


@router.get("/obras", response_model=list[ObraResponse], dependencies=GESTOR_ONLY)
def listar_obras(
    busca: str | None = Query(None),
    incluir_inativas: bool = False,
    db=Depends(get_supabase),
):
    """Lista obras; opcionalmente filtra por termo (nome/cidade/cliente Celesc)."""
    try:
        query = db.table("obras").select("*, clientes(nome)")
        if not incluir_inativas:
            query = query.eq("ativo", True)
        if busca:
            busca = _termo_busca_seguro(busca)
            query = query.or_(f"nome.ilike.%{busca}%,cidade.ilike.%{busca}%,cliente_celesc.ilike.%{busca}%")
        return query.order("nome").execute().data
    except Exception:
        logger.exception("Erro ao listar obras")
        raise HTTPException(status_code=500, detail="Erro ao listar obras.") from None


@router.post("/obras", response_model=ObraResponse, status_code=201, dependencies=GESTOR_ONLY)
def criar_obra(obra: ObraCreate, db=Depends(get_supabase)):
    try:
        dados = _validar_cliente_obra(db, obra)
        resp = db.table("obras").insert(dados).execute()
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao criar obra.")
        return resp.data[0]
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao criar obra")
        raise HTTPException(status_code=500, detail="Erro ao criar obra.") from None


@router.put("/obras/{obra_id}", response_model=ObraResponse, dependencies=GESTOR_ONLY)
def atualizar_obra(obra_id: int, obra: ObraCreate, db=Depends(get_supabase)):
    try:
        _obter_ou_404(db, "obras", obra_id, "Obra")
        dados = _validar_cliente_obra(db, obra)
        resp = db.table("obras").update(dados).eq("id", obra_id).execute()
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao atualizar obra.")
        return resp.data[0]
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao atualizar obra %s", obra_id)
        raise HTTPException(status_code=500, detail="Erro ao atualizar obra.") from None


@router.delete("/obras/{obra_id}", dependencies=GESTOR_ONLY)
def excluir_obra(obra_id: int, db=Depends(get_supabase)):
    """Exclusão lógica: mantém o histórico de O.S íntegro."""
    try:
        _obter_ou_404(db, "obras", obra_id, "Obra")
        usadas = db.table("ordens_servico").select("id").eq("obra_id", obra_id).limit(1).execute()
        if usadas.data:
            db.table("obras").update({"ativo": False}).eq("id", obra_id).execute()
            return {"success": True, "message": "Obra possui O.S vinculadas e foi apenas inativada."}
        db.table("obras").delete().eq("id", obra_id).execute()
        return {"success": True, "message": "Obra excluída com sucesso."}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao excluir obra %s", obra_id)
        raise HTTPException(status_code=500, detail="Erro ao excluir obra.") from None


# ---------------------------------------------------------------------------
# Equipes
# ---------------------------------------------------------------------------


@router.get("/equipes", response_model=list[EquipeResponse], dependencies=GESTOR_ONLY)
def listar_equipes(db=Depends(get_supabase)):
    try:
        equipes = db.table("equipes").select("*").order("nome").execute().data
        if not equipes:
            return []

        # Membros de TODAS as equipes em UMA consulta (era 1 query por equipe).
        vinculos = (
            db.table("equipe_membros")
            .select("id, equipe_id, funcionario_id, lider, funcionarios(nome, cpf)")
            .in_("equipe_id", [e["id"] for e in equipes])
            .order("id")
            .execute()
            .data
        )
        por_equipe: dict[int, list[dict]] = {e["id"]: [] for e in equipes}
        for v in vinculos or []:
            func = v.get("funcionarios") or {}
            por_equipe.setdefault(v["equipe_id"], []).append(
                {
                    "id": v["id"],
                    "funcionario_id": v["funcionario_id"],
                    "nome": func.get("nome"),
                    "lider": v.get("lider", False),
                }
            )

        return [{**e, "membros": por_equipe.get(e["id"], [])} for e in equipes]
    except Exception:
        logger.exception("Erro ao listar equipes")
        raise HTTPException(status_code=500, detail="Erro ao listar equipes.") from None


@router.post("/equipes", response_model=EquipeResponse, status_code=201, dependencies=GESTOR_ONLY)
def criar_equipe(equipe: EquipeCreate, db=Depends(get_supabase)):
    try:
        dup = db.table("equipes").select("id").eq("nome", equipe.nome).execute()
        if dup.data:
            raise HTTPException(status_code=400, detail="Já existe uma equipe com este nome.")
        # Regra: o líder deve fazer parte da equipe.
        if equipe.lider_id is not None and equipe.lider_id not in equipe.membro_ids:
            raise HTTPException(status_code=400, detail="O líder deve ser um membro da equipe.")
        # Valida a existência dos membros ANTES do insert (sem equipe órfã).
        _validar_membros_existem(db, equipe.membro_ids)
        resp = (
            db.table("equipes")
            .insert(
                {
                    "nome": equipe.nome,
                    "numero": equipe.numero,
                    "descricao": equipe.descricao,
                    "ativa": True,
                }
            )
            .execute()
        )
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao criar equipe.")
        nova = resp.data[0]
        _gravar_membros(db, nova["id"], equipe.membro_ids, equipe.lider_id)
        return {**nova, "membros": _membros_da_equipe(db, nova["id"])}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao criar equipe")
        raise HTTPException(status_code=500, detail="Erro ao criar equipe.") from None


@router.put("/equipes/{equipe_id}", response_model=EquipeResponse, dependencies=GESTOR_ONLY)
def atualizar_equipe(equipe_id: int, equipe: EquipeCreate, db=Depends(get_supabase)):
    try:
        _obter_ou_404(db, "equipes", equipe_id, "Equipe")
        if equipe.lider_id is not None and equipe.lider_id not in equipe.membro_ids:
            raise HTTPException(status_code=400, detail="O líder deve ser um membro da equipe.")
        # Valida a existência dos membros antes de alterar a equipe.
        _validar_membros_existem(db, equipe.membro_ids)
        resp = (
            db.table("equipes")
            .update(
                {
                    "nome": equipe.nome,
                    "numero": equipe.numero,
                    "descricao": equipe.descricao,
                }
            )
            .eq("id", equipe_id)
            .execute()
        )
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao atualizar equipe.")
        _gravar_membros(db, equipe_id, equipe.membro_ids, equipe.lider_id)
        return {**resp.data[0], "membros": _membros_da_equipe(db, equipe_id)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao atualizar equipe %s", equipe_id)
        raise HTTPException(status_code=500, detail="Erro ao atualizar equipe.") from None


@router.delete("/equipes/{equipe_id}", dependencies=GESTOR_ONLY)
def excluir_equipe(equipe_id: int, db=Depends(get_supabase)):
    try:
        _obter_ou_404(db, "equipes", equipe_id, "Equipe")
        usadas = db.table("ordens_servico").select("id").eq("equipe_id", equipe_id).limit(1).execute()
        if usadas.data:
            db.table("equipes").update({"ativa": False}).eq("id", equipe_id).execute()
            return {"success": True, "message": "Equipe possui O.S vinculadas e foi apenas desativada."}
        db.table("equipes").delete().eq("id", equipe_id).execute()
        return {"success": True, "message": "Equipe excluída com sucesso."}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao excluir equipe %s", equipe_id)
        raise HTTPException(status_code=500, detail="Erro ao excluir equipe.") from None


# ---------------------------------------------------------------------------
# Produtos (serviços por contrato)
# ---------------------------------------------------------------------------

def _validar_tipo_servico(tipo: str) -> None:
    if tipo not in TIPOS_SERVICO:
        raise HTTPException(
            status_code=422,
            detail=f"Contrato inválido: '{tipo}'. Use 'construcao', 'manutencao' ou 'linha_viva'.",
        )


def _codigo_produto_em_uso(db, valor: str, tipo: str, ignorar_id: int | None = None) -> dict | None:
    """Procura `valor` em `codigo`/`codigo_especial` de OUTRO serviço do MESMO
    contrato — ou de um legado (tipo NULL), que vale para todos os contratos.

    CONTRATOS INDEPENDENTES: cada contrato tem o seu catálogo, portanto o
    mesmo código pode existir em contratos diferentes (registros separados).
    A ambigüidade (código normal de um serviço = especial de outro) só vale
    dentro do mesmo contrato. O próprio registro editado é ignorado.
    """
    for coluna in ("codigo", "codigo_especial"):
        resp = db.table("produtos").select("id, nome, tipo").eq(coluna, valor).execute()
        for linha in resp.data or []:
            if ignorar_id is not None and linha["id"] == ignorar_id:
                continue
            tipo_linha = linha.get("tipo")
            if tipo_linha == tipo or tipo_linha is None:
                return linha
    return None


def _carregar_catalogo_servicos(db) -> tuple[dict[str, list], dict[str, list]]:
    """Carrega o catálogo de serviços em memória para a importação em lote.

    Devolve dois mapas (chave = código, valor = registros):
      - `por_codigo_normal`: apenas registros cujo CÓDIGO NORMAL é a chave
        (usado para decidir entre atualizar/adotar/criar);
      - `por_codigo_qualquer`: registros cujo código normal OU especial é a
        chave (usado para detectar colisões dentro do mesmo contrato).

    Com isso a importação resolve TODAS as linhas sem 4-6 consultas síncronas
    por linha (antes, timeout em arquivos grandes).
    """
    por_normal: dict[str, list] = {}
    por_qualquer: dict[str, list] = {}
    base = db.table("produtos").select("id, codigo, codigo_especial, tipo, ativo, nome").order("id")
    offset = 0
    tamanho = 1000
    while True:
        pagina = base.range(offset, offset + tamanho - 1).execute().data
        if not pagina:
            break
        for p in pagina:
            if p.get("codigo"):
                por_normal.setdefault(str(p["codigo"]), []).append(p)
            for chave in (p.get("codigo"), p.get("codigo_especial")):
                if chave:
                    por_qualquer.setdefault(str(chave), []).append(p)
        if len(pagina) < tamanho:
            break
        offset += tamanho
    return por_normal, por_qualquer


def _servico_em_uso_no_catalogo(lista: list[dict], tipo: str, ignorar_id: int | None) -> dict | None:
    """Equivale a `_codigo_produto_em_uso`, resolvendo contra o catálogo em
    memória (mesmo contrato ou legado; o próprio registro é ignorado)."""
    for linha in lista:
        if ignorar_id is not None and linha["id"] == ignorar_id:
            continue
        tipo_linha = linha.get("tipo")
        if tipo_linha == tipo or tipo_linha is None:
            return linha
    return None


def _validar_codigos_produto(db, produto: ProdutoCreate, produto_id: int | None = None) -> None:
    """Normalização já feita no schema; aqui valida colisões entre os códigos
    (sempre no escopo do contrato do serviço)."""
    codigo = produto.codigo
    codigo_especial = produto.codigo_especial
    if codigo and codigo == codigo_especial:
        raise HTTPException(
            status_code=400,
            detail="O código normal e o código especial devem ser diferentes.",
        )
    for rotulo, valor in (("normal", codigo), ("especial", codigo_especial)):
        if not valor:
            continue
        em_uso = _codigo_produto_em_uso(db, valor, produto.tipo, produto_id)
        if em_uso:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Já existe um serviço ('{em_uso['nome']}') com o código {rotulo} '{valor}' "
                    "neste contrato (ou como legado). Edite o serviço existente para alterá-lo."
                ),
            )


@router.get("/produtos", response_model=list[ProdutoResponse])
def listar_produtos(
    busca: str | None = Query(None, description="Busca por nome ou código (autocompletar)"),
    tipo: str | None = Query(None, description="Filtra pelo contrato (legados sem tipo valem para todos)"),
    db=Depends(get_supabase),
):
    """Lista TODOS os serviços ativos do catálogo.

    O catálogo alimenta o cadastro, o lançamento de serviços na O.S e o
    pacote offline do Modo Campo — por isso NÃO pode ser truncado (era
    limitado a 50 linhas, fazendo serviços sumirem da busca/cadastro). A
    busca por páginas via `range` contorna o teto de linhas por requisição
    do PostgREST/Supabase (padrão de 1000).
    """
    try:
        base = db.table("produtos").select("*").eq("ativo", True)
        if busca:
            termo = _termo_busca_seguro(busca)
            if termo:
                base = base.or_(
                    f"nome.ilike.%{termo}%,codigo.ilike.%{termo}%,codigo_especial.ilike.%{termo}%"
                )
        base = base.order("nome")

        dados: list[dict] = []
        tamanho_pagina = 1000
        offset = 0
        while True:
            pagina = base.range(offset, offset + tamanho_pagina - 1).execute().data
            dados.extend(pagina)
            if len(pagina) < tamanho_pagina:
                break
            offset += tamanho_pagina

        if tipo:
            # Filtro ESTRITO por contrato: cada contrato tem o seu catálogo.
            # Legados (tipo NULL) valem para todos os contratos.
            dados = [p for p in dados if p.get("tipo") is None or p["tipo"] == tipo]
        return dados
    except Exception:
        logger.exception("Erro ao listar produtos")
        raise HTTPException(status_code=500, detail="Erro ao listar serviços.") from None


@router.post("/produtos", response_model=ProdutoResponse, status_code=201, dependencies=GESTOR_ONLY)
def criar_produto(produto: ProdutoCreate, db=Depends(get_supabase)):
    try:
        _validar_tipo_servico(produto.tipo)
        _validar_codigos_produto(db, produto)
        resp = db.table("produtos").insert(produto.model_dump()).execute()
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao criar serviço.")
        return resp.data[0]
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao criar produto")
        raise HTTPException(status_code=500, detail="Erro ao criar serviço.") from None


@router.put("/produtos/{produto_id}", response_model=ProdutoResponse, dependencies=GESTOR_ONLY)
def atualizar_produto(produto_id: int, produto: ProdutoCreate, db=Depends(get_supabase)):
    try:
        _obter_ou_404(db, "produtos", produto_id, "Serviço")
        _validar_tipo_servico(produto.tipo)
        _validar_codigos_produto(db, produto, produto_id=produto_id)
        resp = db.table("produtos").update(produto.model_dump()).eq("id", produto_id).execute()
        if not resp.data:
            raise HTTPException(status_code=500, detail="Falha ao atualizar serviço.")
        return resp.data[0]
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao atualizar produto %s", produto_id)
        raise HTTPException(status_code=500, detail="Erro ao atualizar serviço.") from None


@router.delete("/produtos/{produto_id}", dependencies=GESTOR_ONLY)
def excluir_produto(produto_id: int, db=Depends(get_supabase)):
    try:
        _obter_ou_404(db, "produtos", produto_id, "Serviço")
        usadas = db.table("os_materiais").select("id").eq("produto_id", produto_id).limit(1).execute()
        if usadas.data:
            db.table("produtos").update({"ativo": False}).eq("id", produto_id).execute()
            return {"success": True, "message": "Serviço possui lançamentos e foi apenas inativado."}
        db.table("produtos").delete().eq("id", produto_id).execute()
        return {"success": True, "message": "Serviço excluído com sucesso."}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao excluir produto %s", produto_id)
        raise HTTPException(status_code=500, detail="Erro ao excluir serviço.") from None


# ---------------------------------------------------------------------------
# Importação em lote de serviços (.xlsx) — mesmo padrão do módulo Comprovantes:
# modelo para download + importação com simulação e relatório por linha.
# ---------------------------------------------------------------------------

MAX_IMPORT_SIZE = 10 * 1024 * 1024  # 10 MB

# Ordem e cabeçalho (amigável) das colunas do modelo de importação.
CAMPOS_MODELO_SERVICO = [
    ("nome", "Serviço (descrição)"),
    ("codigo", "Código Normal"),
    ("codigo_especial", "Código Especial"),
    ("unidade", "Unidade"),
    ("preco_unitario", "Qtd USC"),
    ("qtd_usc_especial", "Qtd USC Especial"),
]

# Apelidos de cabeçalho (normalizados) -> nome do campo. Aceita variações
# reais de planilhas (inclusive a listagem oficial de serviços).
ALIASES_COLUNA_SERVICO = {
    "servico": "nome",
    "servico (descricao)": "nome",
    "servico (descrição)": "nome",
    "descricao": "nome",
    "descricao do servico": "nome",
    "descricao do serviço": "nome",
    "nome": "nome",
    "nome do servico": "nome",
    "nome do serviço": "nome",
    "codigo normal": "codigo",
    "codigo": "codigo",
    "codigo sku": "codigo",
    "codigo de barras": "codigo",
    "sku": "codigo",
    "codigo especial": "codigo_especial",
    "codigo usc especial": "codigo_especial",
    "unidade": "unidade",
    "qtd usc": "preco_unitario",
    "usc": "preco_unitario",
    "usc normal": "preco_unitario",
    "quantidade usc": "preco_unitario",
    "preco unitario": "preco_unitario",
    "preço unitário": "preco_unitario",
    "valor": "preco_unitario",
    "qtd usc especial": "qtd_usc_especial",
    "usc especial": "qtd_usc_especial",
    "quantidade usc especial": "qtd_usc_especial",
}


def _normalizar(texto):
    """Minúsculas, sem acentos e com espaços simples (para casar cabeçalhos)."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", texto)


def _parse_numero_servico(valor):
    """Converte número (célula numérica, '0,48', '0.48', '1.500,00') para float.

    Retorna None quando não é um número válido. Vazio -> None (o chamador
    decide o default).
    """
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    texto = str(valor).strip().replace("R$", "").replace(" ", "")
    if not texto:
        return None
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return round(float(texto), 2)
    except ValueError:
        return None


def _planilha_para_servicos(conteudo: bytes) -> tuple[list[dict], int]:
    """Lê o .xlsx e devolve (registros por linha, linhas totalmente vazias ignoradas)."""
    try:
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception:
        logger.exception("Falha ao ler planilha de serviços")
        raise HTTPException(
            status_code=400,
            detail="Não foi possível ler o arquivo. Verifique se é um .xlsx válido.",
        ) from None

    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas or not any(linhas[0]):
        raise HTTPException(status_code=400, detail="A planilha está vazia ou sem cabeçalho.")

    # Mapeia cabeçalho -> índice da coluna.
    colunas = {}
    for idx, valor in enumerate(linhas[0]):
        campo = ALIASES_COLUNA_SERVICO.get(_normalizar(valor))
        if campo and campo not in colunas:
            colunas[campo] = idx

    if "nome" not in colunas:
        raise HTTPException(
            status_code=400,
            detail="Nenhuma coluna reconhecida. Use o modelo baixado em 'Baixar modelo'.",
        )

    registros = []
    vazias = 0
    for num, linha in enumerate(linhas[1:], start=2):
        if all(celula is None or str(celula).strip() == "" for celula in linha):
            vazias += 1
            continue
        registro = {"linha": num}
        for campo, idx in colunas.items():
            if idx < len(linha):
                registro[campo] = linha[idx]
        registros.append(registro)
    return registros, vazias


@router.get("/produtos/modelo", dependencies=GESTOR_ONLY)
def modelo_servicos():
    """Gera e baixa um modelo .xlsx pronto para o cadastro de serviços em lote."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo"

        cabecalhos = [rotulo for _, rotulo in CAMPOS_MODELO_SERVICO]

        fonte_cabecalho = Font(bold=True, color="FFFFFF")
        fill_cabecalho = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        borda = Border(*[Side(style="thin", color="CBD5E1")] * 4)

        for col, rotulo in enumerate(cabecalhos, start=1):
            celula = ws.cell(row=1, column=col, value=rotulo)
            celula.font = fonte_cabecalho
            celula.fill = fill_cabecalho
            celula.alignment = Alignment(horizontal="center", vertical="center")
            celula.border = borda

        exemplos = [
            {
                "nome": "Corte de árvore com risco iminente",
                "codigo": "CRA-01",
                "codigo_especial": "CRA-ESP",
                "unidade": "UN",
                "preco_unitario": 0.48,
                "qtd_usc_especial": 0.67,
            },
            {
                "nome": "Roçada de capoeira",
                "codigo": "ROC-01",
                "codigo_especial": "",
                "unidade": "m²",
                "preco_unitario": 6.66,
                "qtd_usc_especial": "",
            },
        ]

        for i, exemplo in enumerate(exemplos, start=2):
            for col, (campo, _) in enumerate(CAMPOS_MODELO_SERVICO, start=1):
                celula = ws.cell(row=i, column=col, value=exemplo.get(campo, ""))
                celula.border = borda
                celula.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        for col, (_, rotulo) in enumerate(CAMPOS_MODELO_SERVICO, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(18, len(rotulo) + 4)

        ws_instrucoes = wb.create_sheet("Instruções")
        linhas = [
            "INSTRUÇÕES — CADASTRO DE SERVIÇOS EM LOTE",
            "",
            "1. Preencha o arquivo .xlsx com um serviço por linha.",
            "2. A primeira linha (cabeçalho) deve permanecer como está. Não altere ou remova.",
            "3. Coluna 'Serviço (descrição)' é obrigatória em todas as linhas.",
            "4. A importação vale SOMENTE para o CONTRATO escolhido na tela: cada",
            "   contrato (Construção/Manutenção/Linha Viva) tem o seu catálogo.",
            "5. Se o 'Código Normal' já existir NESTE contrato, o serviço é ATUALIZADO;",
            "   se existir só como legado (sem contrato), ele é adotado por este contrato;",
            "   caso contrário, um novo serviço é criado.",
            "6. O MESMO código pode existir em contratos diferentes (são catálogos",
            "   independentes) — importar em um contrato NÃO altera os demais.",
            "7. 'Código Especial' é o código usado quando o serviço é aplicado como USC especial",
            "   (mesma descrição, dois códigos distintos).",
            "8. Números: use vírgula como separador decimal (ex.: 0,48 ou 6,66).",
            "9. Ao finalizar, vá em 'Importar em lote' na aba Serviços e envie este arquivo.",
            "10. A importação pode ser simulada primeiro (prévia) para conferir antes de aplicar.",
            "11. ATENÇÃO: se o código começar com zeros (ex.: 001234), formate a coluna como",
            "    TEXTO no Excel ANTES de digitar — senão o Excel remove os zeros à esquerda.",
            "12. Códigos totalmente numéricos (código de barras) são aceitos e normalizados",
            "    automaticamente (sem decimal no final, ex.: 75012300000000).",
        ]
        for i, texto in enumerate(linhas, start=1):
            ws_instrucoes.cell(row=i, column=1, value=texto)
        ws_instrucoes.column_dimensions["A"].width = 110

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="modelo_servicos.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao gerar modelo de importação de serviços")
        raise HTTPException(status_code=500, detail="Erro ao gerar modelo de importação.") from None


def _mensagem_erro_banco(exc: Exception) -> str:
    """Extrai uma mensagem legível de exceções do Supabase/PostgREST.

    O supabase-py levanta `APIError` (dict-like) com a chave `message` — o
    mesmo formato de qualquer erro HTTP do PostgREST (coluna inexistente,
    violação de unicidade etc.).
    """
    mensagem = getattr(exc, "message", None)
    if not mensagem and exc.args:
        primeiro = exc.args[0]
        if isinstance(primeiro, dict):
            mensagem = primeiro.get("message") or primeiro.get("details")
    if isinstance(mensagem, str) and mensagem.strip():
        return mensagem.strip()
    texto = str(exc).strip()
    return texto or "erro desconhecido no banco de dados"


def _verificar_colunas_produtos(db) -> str | None:
    """Confirma que as colunas do catálogo existem no banco antes de importar.

    Se o schema.sql ainda não foi aplicado no banco, o PostgREST responde 42703
    ("column ... does not exist") a cada consulta com `codigo_especial`. Em vez
    de um 500 genérico, devolvemos a orientação exata de correção.
    """
    try:
        db.table("produtos").select("id, codigo, codigo_especial").limit(1).execute()
        return None
    except Exception as exc:
        logger.warning("Falha ao sondar colunas de produtos: %s", _mensagem_erro_banco(exc))
        mensagem = _mensagem_erro_banco(exc).lower()
        if "codigo_especial" in mensagem or "codigo_servico" in mensagem:
            return (
                "O banco de dados está desatualizado em relação ao schema.sql: falta a coluna "
                "indicada pelo banco. Execute no Supabase (SQL Editor): "
                "ALTER TABLE produtos ADD COLUMN IF NOT EXISTS codigo_especial VARCHAR(50) UNIQUE; "
                "e ALTER TABLE os_materiais ADD COLUMN IF NOT EXISTS codigo_servico VARCHAR(50);"
            )
        return None


@router.post("/produtos/importar", dependencies=GESTOR_ONLY)
def importar_servicos(
    file: UploadFile = File(...),
    tipo: str = Form(..., description="Contrato dos serviços importados: construcao, manutencao ou linha_viva"),
    simular: bool = Form(False),
    db=Depends(get_supabase),
):
    """Cadastra/atualiza serviços em lote a partir de um .xlsx.

    Upsert pelo Código Normal: se o serviço já existe, atualiza os campos da
    linha; senão cria. Com `simular=True` apenas valida e informa o que seria
    feito, sem gravar nada no banco.
    """
    try:
        _validar_tipo_servico(tipo)

        # Fail-fast: se o banco ainda não tem as colunas novas (schema.sql),
        # orienta a correção em vez de falhar por linha / dar 500 genérico.
        orientacao = _verificar_colunas_produtos(db)
        if orientacao:
            raise HTTPException(status_code=400, detail=orientacao)

        filename = (file.filename or "").lower()
        if not filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Formato inválido. Envie um arquivo .xlsx.")

        conteudo = file.file.read(MAX_IMPORT_SIZE + 1)
        if len(conteudo) > MAX_IMPORT_SIZE:
            raise HTTPException(status_code=400, detail="Arquivo muito grande. O tamanho máximo é 10 MB.")
        if len(conteudo) == 0:
            raise HTTPException(status_code=400, detail="Arquivo vazio.")

        registros, vazias = _planilha_para_servicos(conteudo)
        if not registros:
            raise HTTPException(status_code=400, detail="Nenhuma linha com dados foi encontrada na planilha.")

        criados = 0
        atualizados = 0
        ignoradas = vazias
        erros = []
        # Namespace de códigos do PRÓPRIO arquivo (linha -> código), para
        # detectar duplicidades/colisões internas sem depender do banco.
        visto: dict[str, int] = {}

        # Catálogo em memória: resolve alvo/colisão de TODAS as linhas sem
        # 4-6 consultas síncronas por linha (timeout em arquivos grandes).
        catalogo_normal, catalogo_qualquer = _carregar_catalogo_servicos(db)

        # Ações acumuladas para aplicar EM LOTE ao final (inserts agrupados).
        pendentes_insert: list[tuple[dict, int]] = []  # (campos, linha)
        pendentes_update: list[tuple[dict, int, int]] = []  # (campos, id, linha)

        for registro in registros:
            num = registro["linha"]
            nome = _normalizar_texto_livre(registro.get("nome"))
            if not nome:
                erros.append({"linha": num, "mensagem": "Informe o nome do serviço (coluna 'Serviço')."})
                continue

            codigo = _texto_ou_none(registro.get("codigo"))
            codigo_especial = _texto_ou_none(registro.get("codigo_especial"))
            unidade = _normalizar_texto_livre(registro.get("unidade")) or "UN"

            # Campos numéricos: vazio -> 0; texto inválido ou negativo -> erro.
            valores = {}
            numeros_ok = True
            for campo, rotulo in (
                ("preco_unitario", "Qtd USC"),
                ("qtd_usc_especial", "Qtd USC Especial"),
            ):
                bruto = registro.get(campo)
                if bruto is None or str(bruto).strip() == "":
                    valores[campo] = 0.0
                    continue
                valor = _parse_numero_servico(bruto)
                if valor is None or valor < 0:
                    erros.append({"linha": num, "mensagem": f"'{rotulo}' inválida na linha."})
                    numeros_ok = False
                    break
                valores[campo] = valor
            if not numeros_ok:
                continue

            if codigo and codigo == codigo_especial:
                erros.append({"linha": num, "mensagem": "O código normal e o código especial devem ser diferentes."})
                continue

            # Duplicidade/colisão dentro do próprio arquivo.
            conflito_linha = None
            for valor in (codigo, codigo_especial):
                if valor and valor in visto and visto[valor] != num:
                    conflito_linha = visto[valor]
                    break
            if conflito_linha:
                erros.append(
                    {
                        "linha": num,
                        "mensagem": f"Código '{valor}' já utilizado na linha {conflito_linha} do arquivo.",
                    }
                )
                continue

            # Upsert POR CONTRATO (resolvido no catálogo em memória):
            # 1) serviço do MESMO contrato com o código  -> atualiza;
            # 2) legado (tipo NULL) com o código         -> é ADOTADO pelo
            #    contrato importado (vira cadastro do contrato);
            # 3) nenhum dos dois                         -> cria novo.
            # Nunca altera serviço de OUTRO contrato com o mesmo código.
            alvo = None
            adotando_legado = False
            if codigo:
                candidatos = catalogo_normal.get(str(codigo), [])
                for linha in candidatos:
                    if linha.get("tipo") == tipo:
                        alvo = linha
                        break
                if alvo is None:
                    for linha in candidatos:
                        if linha.get("tipo") is None:
                            alvo = linha
                            adotando_legado = True
                            break
            ignorar_id = alvo["id"] if alvo else None

            # Colisão no MESMO contrato (ou contra legado não adotado).
            colisao = None
            for valor in (codigo, codigo_especial):
                if not valor:
                    continue
                em_uso = _servico_em_uso_no_catalogo(catalogo_qualquer.get(str(valor), []), tipo, ignorar_id)
                if em_uso:
                    colisao = em_uso
                    break
            if colisao:
                erros.append(
                    {
                        "linha": num,
                        "mensagem": f"Já existe o serviço '{colisao['nome']}' com o código '{valor}' "
                        "(cadastre com outro código ou edite o serviço existente).",
                    }
                )
                continue

            campos = {
                "nome": nome,
                "codigo": codigo,
                "codigo_especial": codigo_especial,
                "unidade": unidade,
                "preco_unitario": valores["preco_unitario"],
                "qtd_usc_especial": valores["qtd_usc_especial"],
            }
            if adotando_legado:
                campos["tipo"] = tipo

            if simular:
                if alvo:
                    atualizados += 1
                else:
                    criados += 1
            elif alvo:
                if alvo.get("ativo") is False:
                    campos["ativo"] = True
                pendentes_update.append((campos, alvo["id"], num))
            else:
                pendentes_insert.append(({**campos, "tipo": tipo, "ativo": True}, num))

            # Registra os códigos da linha no namespace do arquivo (apenas
            # linhas aceitas — inclusive na simulação).
            for valor in (codigo, codigo_especial):
                if valor:
                    visto[valor] = num

        # ---- Aplicação (fora do loop; só quando não é simulação) ----------
        if not simular:
            # Atualizações: uma por registro (não há batch de UPDATE por id).
            for campos, alvo_id, num in pendentes_update:
                try:
                    db.table("produtos").update(campos).eq("id", alvo_id).execute()
                    atualizados += 1
                except HTTPException:
                    raise
                except Exception as exc:
                    logger.exception("Erro de banco ao importar linha %d", num)
                    causa = _mensagem_erro_banco(exc)
                    if "value too long for type character varying" in causa.lower():
                        causa = (
                            "Texto muito longo para o banco (coluna ainda com limite de 255). "
                            "Encurte a descrição ou amplie a coluna: "
                            "ALTER TABLE produtos ALTER COLUMN nome TYPE TEXT;"
                        )
                    erros.append({"linha": num, "mensagem": f"Falha no banco: {causa}"})

            # Novos serviços: inserts em LOTE de 100 (era 1 chamada por linha).
            for inicio in range(0, len(pendentes_insert), 100):
                bloco = pendentes_insert[inicio:inicio + 100]
                try:
                    resp = db.table("produtos").insert([campos for campos, _ in bloco]).execute()
                    criados += len(resp.data or [])
                except HTTPException:
                    raise
                except Exception:
                    # Lote inteiro recusado: tenta linha a linha para isolar a
                    # real e mantê-la no relatório (sem derrubar a importação).
                    logger.exception("Falha em lote de insert (%d linhas); isolando linhas.", len(bloco))
                    for campos, num in bloco:
                        try:
                            r2 = db.table("produtos").insert(campos).execute()
                            if r2.data:
                                criados += 1
                        except HTTPException:
                            raise
                        except Exception as exc:
                            causa = _mensagem_erro_banco(exc)
                            if "value too long for type character varying" in causa.lower():
                                causa = (
                                    "Texto muito longo para o banco (coluna ainda com limite de 255). "
                                    "Encurte a descrição ou amplie a coluna: "
                                    "ALTER TABLE produtos ALTER COLUMN nome TYPE TEXT;"
                                )
                            erros.append({"linha": num, "mensagem": f"Falha no banco: {causa}"})

        return {
            "importados": criados + atualizados,
            "criados": criados,
            "atualizados": atualizados,
            "erros": erros,
            "total": len(registros),
            "ignoradas": ignoradas,
            "simular": simular,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Erro ao importar serviços em lote")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao importar serviços. Detalhe: {_mensagem_erro_banco(exc)}",
        ) from None


def _texto_ou_none(valor) -> str | None:
    """Converte célula para texto, normalizando números.

    Células numéricas do Excel (códigos de barras/SKU longos) chegam como
    float — ex.: 75012300000000.0. Aqui viram texto sem o sufixo '.0'
    ('75012300000000'), preservando a comparabilidade com a bipagem.
    """
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip()
    return texto or None


def _normalizar_texto_livre(valor) -> str:
    """Strip simples de células de texto (sem normalizar caixa/conteúdo)."""
    if valor is None:
        return ""
    return str(valor).strip()
