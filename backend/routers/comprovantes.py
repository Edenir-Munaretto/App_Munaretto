import io
import logging
import re
import unicodedata
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Query, Depends, UploadFile, File, Form
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field
from supabase_client import get_supabase
from auth import get_current_user, require_permisao

router = APIRouter(dependencies=[Depends(require_permisao("comprovantes"))])

logger = logging.getLogger(__name__)

MAX_IMPORT_SIZE = 10 * 1024 * 1024  # 10 MB

# Ordem e cabeçalho (amigável) das colunas do modelo de importação.
CAMPOS_MODELO = [
    ("tipo_documento", "Tipo do Documento"),
    ("nome", "Nome"),
    ("cnpj", "CNPJ/CPF"),
    ("numero_nf", "Número NF"),
    ("data_emissao", "Data de Emissão"),
    ("data_vencimento", "Data de Vencimento"),
    ("data_pagamento", "Data de Pagamento"),
    ("descricao", "Descrição"),
    ("valor_total", "Valor Total"),
    ("base_calculo", "Base de Cálculo"),
    ("valor_inss", "Valor INSS"),
    ("valor_iss", "Valor ISS"),
    ("valor_liquido", "Valor Líquido"),
    ("valor_pago", "Valor Pago"),
    ("valor_juros", "Valor Juros"),
    ("local_servico", "Local de Serviço"),
    ("forma_pagamento", "Forma de Pagamento"),
]

TIPOS_VALIDOS = ["Nota Fiscal", "Boleto", "Pix", "Diversas", "Aluguel", "Imposto"]

# Apelidos de cabeçalho (normalizados) -> nome do campo no banco.
ALIASES_COLUNA = {
    "tipo documento": "tipo_documento",
    "tipo do documento": "tipo_documento",
    "tipo_documento": "tipo_documento",
    "documento": "tipo_documento",
    "nome": "nome",
    "nome do fornecedor": "nome",
    "beneficiario": "nome",
    "fornecedor": "nome",
    "credor": "nome",
    "cnpj": "cnpj",
    "cnpj cpf": "cnpj",
    "cnpj/cpf": "cnpj",
    "cpf": "cnpj",
    "cpf cnpj": "cnpj",
    "cpf/cnpj": "cnpj",
    "numero nf": "numero_nf",
    "numero_nf": "numero_nf",
    "nf": "numero_nf",
    "nota fiscal numero": "numero_nf",
    "data emissao": "data_emissao",
    "data de emissao": "data_emissao",
    "data_emissao": "data_emissao",
    "emissao": "data_emissao",
    "data vencimento": "data_vencimento",
    "data de vencimento": "data_vencimento",
    "data_vencimento": "data_vencimento",
    "vencimento": "data_vencimento",
    "data pagamento": "data_pagamento",
    "data de pagamento": "data_pagamento",
    "data_pagamento": "data_pagamento",
    "pagamento": "data_pagamento",
    "descricao": "descricao",
    "descricao do pagamento": "descricao",
    "descricao da despesa": "descricao",
    "historico": "descricao",
    "observacao": "descricao",
    "valor total": "valor_total",
    "valor_total": "valor_total",
    "base calculo": "base_calculo",
    "base de calculo": "base_calculo",
    "base_calculo": "base_calculo",
    "base": "base_calculo",
    "valor inss": "valor_inss",
    "valor_inss": "valor_inss",
    "inss": "valor_inss",
    "valor iss": "valor_iss",
    "valor_iss": "valor_iss",
    "iss": "valor_iss",
    "valor liquido": "valor_liquido",
    "valor_liquido": "valor_liquido",
    "liquido": "valor_liquido",
    "valor pago": "valor_pago",
    "valor_pago": "valor_pago",
    "valor": "valor_pago",
    "valor juros": "valor_juros",
    "valor_juros": "valor_juros",
    "juros": "valor_juros",
    "local servico": "local_servico",
    "local de servico": "local_servico",
    "local_servico": "local_servico",
    "forma pagamento": "forma_pagamento",
    "forma de pagamento": "forma_pagamento",
    "forma_pagamento": "forma_pagamento",
}


def _normalizar(texto):
    """Minúsculas, sem acentos e com espaços simples (para casar cabeçalhos)."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", texto)


def _parse_data(valor):
    """Converte data (datetime, dd/mm/yyyy ou yyyy-mm-dd) para yyyy-mm-dd."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    if not texto:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_numero(valor):
    """Converte número (float, '1.500,00', 'R$ 1.500,00') para float."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return float(valor)
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    texto = str(valor).strip().replace("R$", "").replace(" ", "")
    if not texto:
        return None
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(".", "")
    try:
        return round(float(texto), 2)
    except ValueError:
        return None


def _normalizar_tipo(valor):
    """Casa texto digitado na planilha com os tipos válidos do sistema."""
    if valor is None:
        return None
    norm = _normalizar(valor)
    if not norm:
        return None
    for tipo in TIPOS_VALIDOS:
        if norm == _normalizar(tipo):
            return tipo
    if norm in ("nota", "nf", "nfs", "nfe"):
        return "Nota Fiscal"
    return None

class ComprovanteCreate(BaseModel):
    tipo_documento: str = Field(..., description="Tipo do documento: Nota Fiscal, Boleto, Pix, Diversas, Aluguel, Imposto")
    
    # Campos Nota Fiscal
    numero_nf: Optional[str] = None
    data_emissao: Optional[str] = None
    nome: Optional[str] = None
    cnpj: Optional[str] = None
    local_servico: Optional[str] = None
    valor_total: Optional[float] = 0.0
    base_calculo: Optional[float] = 0.0
    valor_inss: Optional[float] = 0.0
    valor_iss: Optional[float] = 0.0
    valor_liquido: Optional[float] = 0.0
    
    # Outros tipos (Boleto, Pix, Diversas, Aluguel)
    data_pagamento: Optional[str] = None
    data_vencimento: Optional[str] = None
    descricao: Optional[str] = None
    forma_pagamento: Optional[str] = None # "boleto", "dda", "pix"
    valor_pago: Optional[float] = 0.0
    valor_juros: Optional[float] = 0.0

class ComprovanteResponse(ComprovanteCreate):
    id: int
    data_registro: str

@router.get("/", response_model=List[ComprovanteResponse])
def listar_comprovantes(
    ordenar_por: str = Query("data_registro", description="Campo de ordenação: data_registro, data_pagamento ou data_emissao"),
    tipo_documento: Optional[str] = Query(None, description="Filtrar por tipo: Nota Fiscal, Boleto, Pix, Diversas, Aluguel, Imposto"),
    data_inicio: Optional[str] = Query(None, description="Data início do filtro (YYYY-MM-DD) — filtra data_emissao ou data_pagamento"),
    data_fim: Optional[str] = Query(None, description="Data fim do filtro (YYYY-MM-DD)"),
    limit: Optional[int] = Query(None, ge=1, le=10000, description="Máximo de registros a retornar (paginação)"),
    offset: Optional[int] = Query(0, ge=0, description="Registros a pular (paginação)"),
    db = Depends(get_supabase),
):
    """Lista lançamentos de comprovantes. Suporta filtro por tipo e período.

    O Supabase limita cada requisição a 1000 linhas, então a listagem é
    paginada internamente (em blocos de 1000) e retorna o resultado completo,
    salvo quando `limit` é informado (paginação explícita).
    Ordenação decrescente (mais recente primeiro) por `data_registro`,
    `data_pagamento` ou `data_emissao`.
    """
    campo_ordem = ordenar_por if ordenar_por in ("data_registro", "data_pagamento", "data_emissao") else "data_registro"

    # Validações antecipadas (fora do try para não virarem erro 500)
    if tipo_documento and tipo_documento not in TIPOS_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de documento inválido. Valores válidos: {', '.join(TIPOS_VALIDOS)}.",
        )
    if data_inicio:
        d = _parse_data(data_inicio)
        if d is None:
            raise HTTPException(status_code=400, detail="data_inicio inválida. Use YYYY-MM-DD (ou dd/mm/aaaa).")
        data_inicio = d
    if data_fim:
        d = _parse_data(data_fim)
        if d is None:
            raise HTTPException(status_code=400, detail="data_fim inválida. Use YYYY-MM-DD (ou dd/mm/aaaa).")
        data_fim = d

    try:
        todos = []
        offset_atual = offset
        bloco = 1000
        while True:
            # Se houver limit, respeita o teto final da resposta
            teto = None
            if limit is not None:
                teto = offset + limit
                if offset_atual >= teto:
                    break
                tamanho_bloco = min(bloco, teto - offset_atual)
            else:
                tamanho_bloco = bloco

            query = (
                db.table("comprovantes")
                .select("*")
                .order(campo_ordem, desc=True)
                .range(offset_atual, offset_atual + tamanho_bloco - 1)
            )
            # Filtro por tipo de documento
            if tipo_documento:
                query = query.eq("tipo_documento", tipo_documento)
            # Filtro por período: casa data_emissao OU data_pagamento dentro do intervalo
            if data_inicio or data_fim:
                condicoes = []
                for campo in ("data_emissao", "data_pagamento"):
                    partes = []
                    if data_inicio:
                        partes.append(f"{campo}.gte.{data_inicio}")
                    if data_fim:
                        partes.append(f"{campo}.lte.{data_fim}")
                    condicoes.append("and(" + ",".join(partes) + ")")
                query = query.or_(",".join(condicoes))

            response = query.execute()
            if not response.data:
                break
            todos.extend(response.data)
            if len(response.data) < tamanho_bloco:
                break
            offset_atual += tamanho_bloco
        return todos
    except Exception as e:
        logger.exception("Erro ao buscar comprovantes")
        raise HTTPException(status_code=500, detail="Erro ao buscar comprovantes")


@router.get("/exportar")
def exportar_comprovantes(
    ordenar_por: str = Query("data_registro", description="Campo de ordenação: data_registro, data_pagamento ou data_emissao"),
    tipo_documento: Optional[str] = Query(None, description="Filtrar por tipo: Nota Fiscal, Boleto, Pix, Diversas, Aluguel, Imposto"),
    data_inicio: Optional[str] = Query(None, description="Data início do filtro (YYYY-MM-DD)"),
    data_fim: Optional[str] = Query(None, description="Data fim do filtro (YYYY-MM-DD)"),
    db = Depends(get_supabase),
):
    """Exporta os comprovantes filtrados para um arquivo .xlsx (openpyxl)."""
    # Reutiliza a mesma listagem com paginação ampla (sem limit = retorna tudo)
    registros = listar_comprovantes(
        ordenar_por=ordenar_por,
        tipo_documento=tipo_documento,
        data_inicio=data_inicio,
        data_fim=data_fim,
        limit=None,
        offset=0,
        db=db,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Comprovantes"

    colunas = [
        ("tipo_documento", "Tipo do Documento"),
        ("nome", "Nome"),
        ("cnpj", "CNPJ/CPF"),
        ("numero_nf", "Número NF"),
        ("data_emissao", "Data de Emissão"),
        ("data_vencimento", "Data de Vencimento"),
        ("data_pagamento", "Data de Pagamento"),
        ("descricao", "Descrição"),
        ("valor_total", "Valor Total"),
        ("base_calculo", "Base de Cálculo"),
        ("valor_inss", "Valor INSS"),
        ("valor_iss", "Valor ISS"),
        ("valor_liquido", "Valor Líquido"),
        ("valor_pago", "Valor Pago"),
        ("valor_juros", "Valor Juros"),
        ("local_servico", "Local de Serviço"),
        ("forma_pagamento", "Forma de Pagamento"),
        ("data_registro", "Data de Registro"),
    ]

    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    fill_cabecalho = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    borda = Border(*[Side(style="thin", color="CBD5E1")] * 4)

    for col, (_, rotulo) in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col, value=rotulo)
        celula.font = fonte_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    for i, reg in enumerate(registros, start=2):
        for col, (campo, _) in enumerate(colunas, start=1):
            celula = ws.cell(row=i, column=col, value=reg.get(campo))
            celula.border = borda

    ws.freeze_panes = "A2"
    for col, (_, rotulo) in enumerate(colunas, start=1):
        letra = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letra].width = max(16, len(rotulo) + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="comprovantes.xlsx"'},
    )


@router.get("/modelo")
def baixar_modelo_importacao():
    """Gera e baixa um modelo .xlsx pronto para preenchimento e importação em lote."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo"

        cabecalhos = [rotulo for _, rotulo in CAMPOS_MODELO]

        preenchimento = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        fonte_cabecalho = Font(bold=True, color="FFFFFF")
        fill_cabecalho = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        borda = Border(*[Side(style="thin", color="CBD5E1")] * 4)

        for col, rotulo in enumerate(cabecalhos, start=1):
            celula = ws.cell(row=1, column=col, value=rotulo)
            celula.font = fonte_cabecalho
            celula.fill = fill_cabecalho
            celula.alignment = Alignment(horizontal="center", vertical="center")
            celula.border = borda

        exemplos = [
            {
                "tipo_documento": "Boleto",
                "nome": "Concessionária de Energia",
                "cnpj": "00.000.000/0001-00",
                "numero_nf": "",
                "data_emissao": "01/08/2026",
                "data_vencimento": "15/08/2026",
                "data_pagamento": "16/08/2026",
                "descricao": "Conta de energia - unidade São Paulo",
                "valor_total": "1500,00",
                "base_calculo": "",
                "valor_inss": "",
                "valor_iss": "",
                "valor_liquido": "",
                "valor_pago": "1500,00",
                "valor_juros": "0,00",
                "local_servico": "",
                "forma_pagamento": "boleto",
            },
            {
                "tipo_documento": "Nota Fiscal",
                "nome": "Fornecedor de Material Ltda",
                "cnpj": "12.345.678/0001-90",
                "numero_nf": "12345",
                "data_emissao": "10/08/2026",
                "data_vencimento": "",
                "data_pagamento": "",
                "descricao": "",
                "valor_total": "5000,00",
                "base_calculo": "5000,00",
                "valor_inss": "550,00",
                "valor_iss": "250,00",
                "valor_liquido": "4200,00",
                "valor_pago": "",
                "valor_juros": "",
                "local_servico": "São Paulo/SP",
                "forma_pagamento": "",
            },
        ]

        for i, exemplo in enumerate(exemplos, start=2):
            for col, (campo, _) in enumerate(CAMPOS_MODELO, start=1):
                celula = ws.cell(row=i, column=col, value=exemplo.get(campo, ""))
                celula.border = borda
                celula.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        for col, (_, rotulo) in enumerate(CAMPOS_MODELO, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(18, len(rotulo) + 4)

        ws_instrucoes = wb.create_sheet("Instruções")
        linhas = [
            "INSTRUÇÕES DE IMPORTACAO EM LOTE",
            "",
            "1. Preencha o arquivo .xlsx com um lançamento por linha.",
            "2. A primeira linha (cabeçalho) deve permanecer como está. Não altere ou remova.",
            "3. Preencha apenas as colunas que fizerem sentido. Colunas vazias são ignoradas.",
            "4. Datas: use o formato dd/mm/aaaa (ex.: 15/08/2026).",
            "5. Valores: use vírgula como separador decimal (ex.: 1.500,00 ou 1500,00).",
            "6. Tipo do Documento aceita: Nota Fiscal, Boleto, Pix, Diversas, Aluguel, Imposto.",
            "7. Forma de Pagamento aceita: boleto, dda ou pix.",
            "8. Para Nota Fiscal, os campos Nome, CNPJ e Data de Emissão são obrigatórios.",
            "9. Para os demais tipos, são obrigatórios: Descrição e Valor Pago.",
            "10. Ao finalizar, vá em 'Importar' na tela de Comprovantes e envie este arquivo.",
        ]
        for i, texto in enumerate(linhas, start=1):
            ws_instrucoes.cell(row=i, column=1, value=texto)
        ws_instrucoes.column_dimensions["A"].width = 100

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="modelo_comprovantes.xlsx"'},
        )
    except Exception as e:
        logger.exception("Erro ao gerar modelo de importação")
        raise HTTPException(status_code=500, detail="Erro ao gerar modelo de importação.")

@router.post("/importar")
def importar_planilha(
    file: UploadFile = File(...),
    simular: bool = Form(False),
    db = Depends(get_supabase),
):
    """Importa comprovantes/boletos em lote a partir de uma planilha .xlsx.

    Com `simular=True` apenas valida e informa o que seria importado, sem
    gravar nada no banco (útil para testar antes de aplicar).
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Formato inválido. Envie um arquivo .xlsx.")

    conteudo = file.file.read(MAX_IMPORT_SIZE + 1)
    if len(conteudo) > MAX_IMPORT_SIZE:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. O tamanho máximo é 10 MB.")
    if len(conteudo) == 0:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")

    try:
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception as e:
        logger.exception("Falha ao ler planilha")
        raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo. Verifique se é um .xlsx válido.")

    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas or not any(linhas[0]):
        raise HTTPException(status_code=400, detail="A planilha está vazia ou sem cabeçalho.")

    # Mapeia cabeçalho -> índice da coluna
    colunas = {}
    for idx, valor in enumerate(linhas[0]):
        campo = ALIASES_COLUNA.get(_normalizar(valor))
        if campo and campo not in colunas:
            colunas[campo] = idx

    if not colunas:
        raise HTTPException(
            status_code=400,
            detail="Nenhuma coluna reconhecida. Use o modelo baixado em 'Baixar modelo'.",
        )

    dados = linhas[1:]
    importados = 0
    erros = []
    total = 0
    ignoradas = 0

    for num, linha in enumerate(dados, start=2):
        # Linhas totalmente vazias são contadas e ignoradas (aparecem no relatório)
        if all(celula is None or str(celula).strip() == "" for celula in linha):
            ignoradas += 1
            continue
        total += 1

        payload = {}
        for campo, idx in colunas.items():
            if idx < len(linha):
                payload[campo] = linha[idx]

        tipo = _normalizar_tipo(payload.get("tipo_documento"))
        if not tipo:
            erros.append({"linha": num, "mensagem": "Tipo do documento inválido ou ausente."})
            continue
        payload["tipo_documento"] = tipo

        for campo_data in ("data_emissao", "data_vencimento", "data_pagamento"):
            if campo_data in payload:
                payload[campo_data] = _parse_data(payload[campo_data])

        for campo_num in (
            "valor_total", "base_calculo", "valor_inss", "valor_iss",
            "valor_liquido", "valor_pago", "valor_juros",
        ):
            if campo_num in payload:
                payload[campo_num] = _parse_numero(payload[campo_num])

        if "forma_pagamento" in payload:
            forma = _normalizar(payload["forma_pagamento"])
            if forma in ("boleto", "dda", "pix"):
                payload["forma_pagamento"] = forma
            else:
                payload["forma_pagamento"] = None

        for chave, valor in list(payload.items()):
            if valor == "":
                payload[chave] = None

        # Validações obrigatórias por tipo
        if tipo == "Nota Fiscal":
            faltantes = [c for c in ("nome", "cnpj", "data_emissao") if not payload.get(c)]
            if faltantes:
                erros.append({"linha": num, "mensagem": f"Nota Fiscal exige: {', '.join(faltantes)}."})
                continue
        else:
            faltantes = [c for c in ("descricao", "valor_pago") if payload.get(c) is None or payload.get(c) == 0]
            if faltantes:
                erros.append({"linha": num, "mensagem": f"{tipo} exige: {', '.join(faltantes)}."})
                continue

        try:
            if simular:
                importados += 1
            else:
                db.table("comprovantes").insert(payload).execute()
                importados += 1
        except Exception as e:
            logger.exception("Erro ao importar linha %d", num)
            erros.append({"linha": num, "mensagem": "Falha ao salvar no banco."})

    if importados == 0 and not erros and ignoradas == 0:
        raise HTTPException(status_code=400, detail="Nenhuma linha com dados foi encontrada na planilha.")

    return {
        "importados": importados,
        "erros": erros,
        "total": total,
        "ignoradas": ignoradas,
        "simular": simular,
    }

@router.get("/{comprovante_id}", response_model=ComprovanteResponse)
def buscar_comprovante(comprovante_id: int, db = Depends(get_supabase)):
    """Busca um comprovante específico pelo ID."""
    try:
        response = db.table("comprovantes").select("*").eq("id", comprovante_id).execute()
        if not response.data:
            raise HTTPException(status_code=404, detail="Comprovante não encontrado.")
        return response.data[0]
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Erro ao buscar comprovante")
        raise HTTPException(status_code=500, detail="Erro ao buscar comprovante")
@router.post("/", response_model=ComprovanteResponse, status_code=201)
def criar_comprovante(comprovante: ComprovanteCreate, db = Depends(get_supabase)):
    """Cria um novo lançamento de comprovante."""
    try:
        payload = comprovante.model_dump()
        # Higieniza strings vazias para None, evitando erros em campos opcionais e do tipo data
        for key, value in list(payload.items()):
            if value == "":
                payload[key] = None
        response = db.table("comprovantes").insert(payload).execute()
        if not response.data:
            raise HTTPException(status_code=500, detail="Falha ao salvar comprovante.")
        return response.data[0]
    except Exception as e:
        logger.exception("Erro ao criar comprovante")
        raise HTTPException(status_code=500, detail="Erro ao criar comprovante")
@router.put("/{comprovante_id}", response_model=ComprovanteResponse)
def atualizar_comprovante(comprovante_id: int, comprovante: ComprovanteCreate, db = Depends(get_supabase)):
    """Atualiza um comprovante existente."""
    try:
        check = db.table("comprovantes").select("id").eq("id", comprovante_id).execute()
        if not check.data:
            raise HTTPException(status_code=404, detail="Comprovante não encontrado.")

        payload = comprovante.model_dump()
        # Higieniza strings vazias para None, evitando erros em campos opcionais e do tipo data
        for key, value in list(payload.items()):
            if value == "":
                payload[key] = None
        response = db.table("comprovantes").update(payload).eq("id", comprovante_id).execute()
        if not response.data:
            raise HTTPException(status_code=500, detail="Falha ao atualizar comprovante.")
        return response.data[0]
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Erro ao atualizar comprovante")
        raise HTTPException(status_code=500, detail="Erro ao atualizar comprovante")
@router.delete("/{comprovante_id}")
def excluir_comprovante(comprovante_id: int, db = Depends(get_supabase)):
    """Exclui um comprovante do sistema."""
    try:
        check = db.table("comprovantes").select("id").eq("id", comprovante_id).execute()
        if not check.data:
            raise HTTPException(status_code=404, detail="Comprovante não encontrado.")

        db.table("comprovantes").delete().eq("id", comprovante_id).execute()
        return {"status": "success", "message": "Comprovante excluído com sucesso."}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Erro ao excluir comprovante")
        raise HTTPException(status_code=500, detail="Erro ao excluir comprovante")