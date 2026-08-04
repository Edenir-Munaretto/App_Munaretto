import os
import shutil
import webbrowser
from datetime import datetime, timedelta
from docxtpl import DocxTemplate
from fpdf import FPDF  # Certifique-se de que é a fpdf2
import tempfile
from num2words import num2words
import os
from pathlib import Path 
import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 1. CONFIGURAÇÕES INICIAIS
TEMPLATES_DIR = "templates"
OUTPUT_DIR = str(os.path.join(os.path.expanduser("~"), "Downloads"))

def garantir_pastas():
    """Cria as pastas se não existirem."""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

# 2. FUNÇÃO QUE A GUI (TKINTER) PRECISA (LINHA 600 DO SEU ERRO)
def get_templates():
    """Retorna os modelos disponíveis para a lista de seleção da interface."""
    garantir_pastas()
    templates = {}
    if not os.path.exists(TEMPLATES_DIR):
        return templates
    for arquivo in sorted(os.listdir(TEMPLATES_DIR)):
        if arquivo.lower().endswith(".docx"):
            nome_limpo = os.path.splitext(arquivo)[0]
            templates[nome_limpo] = {
                "nome": nome_limpo.replace("_", " ").title(),
                "docx_path": os.path.join(TEMPLATES_DIR, arquivo),
                "template": nome_limpo # Mantém compatibilidade com versões antigas
            }
    return templates
def valor_por_extenso(valor_str):
    try:
        
        limpo = valor_str.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        valor_float = float(limpo)
        
        # 2. Converte para extenso em português e no formato de moeda
        extenso = num2words(valor_float, lang='pt_BR', to='currency')
        return f"({extenso.capitalize()})"
    except Exception as e:
        print(f"Erro ao converter extenso: {e}")
        return ""

# 3. FUNÇÕES DE APOIO
def criar_contexto_cliente(cliente_info):
    """Mapeia os dados do banco de dados para o documento."""
    # Se o cliente_info vier do banco como uma tupla/lista
    hoje = datetime.now()
    data_60_dias = hoje + timedelta(days=60)
    
    return {
        "id": cliente_info[0] if len(cliente_info) > 0 else "",
        "nome": cliente_info[1] if len(cliente_info) > 1 else "",
        "cpf_cnpj": cliente_info[2] if len(cliente_info) > 2 else "",
        "endereco": cliente_info[3] if len(cliente_info) > 3 else "",
        "cidade": cliente_info[4] if len(cliente_info) > 4 else "",
        "cep": cliente_info[5] if len(cliente_info) > 5 else "",
        "nota_ps": cliente_info[6] if len(cliente_info) > 6 else "",
        "valor_de_devolucao": cliente_info[8] if len(cliente_info) > 8 else "0,00",
        "valor_devolucao_extenso": valor_por_extenso(cliente_info[8]) if len(cliente_info) > 8 else "",
        "valor_da_obra": cliente_info[7] if len(cliente_info) > 7 else "0,00",
        "valor_extenso": valor_por_extenso(cliente_info[7]) if len(cliente_info) > 7 else "",
        "data": hoje.strftime("%d/%m/%Y"),
        "data_fim": data_60_dias.strftime("%d/%m/%Y")
    }

def abrir_no_navegador(caminho_arquivo):
    """Abre o arquivo automaticamente."""
    if caminho_arquivo and os.path.exists(caminho_arquivo):
        abs_path = os.path.abspath(caminho_arquivo)
        webbrowser.open(f"file://{abs_path}")

# 4. GERAÇÃO DE WORD
def gerar_documento_word(cliente_info, tipo_documento, saida_dir=OUTPUT_DIR):
    try:
        garantir_pastas()
        contexto = criar_contexto_cliente(cliente_info)
        template_path = os.path.join(TEMPLATES_DIR, f"{tipo_documento}.docx")
        
        if not os.path.exists(template_path):
            return None

        doc = DocxTemplate(template_path)
        doc.render(contexto)

        # Sanitiza o nome e limita a 50 caracteres para evitar erros de caminho muito longo (Windows MAX_PATH)
        nome_limpo = "".join([c for c in str(contexto['nome']) if c.isalnum() or c in (' ', '_')]).strip()
        nome_curto = nome_limpo[:50].replace(' ', '_')
        nome_arq = f"{nome_curto}_{datetime.now().strftime('%H%M%S')}.docx"
        caminho = os.path.join(saida_dir, nome_arq)
        doc.save(caminho)
        return caminho
    except Exception as e:
        print(f"Erro Word: {e}")
        return None

# 5. GERAÇÃO DE PDF
def gerar_documento_pdf(cliente_info, tipo_documento):
    """Gera um PDF que é a cópia exata do Word preenchido."""
    try:
        # Inicializa o COM para a thread atual (obrigatório para executáveis)
        pythoncom.CoInitialize()
        
        temp_dir = tempfile.gettempdir()
        # 1. Primeiro geramos o Word normalmente
        caminho_word = gerar_documento_word(cliente_info, tipo_documento, saida_dir=temp_dir)
        
        if not caminho_word:
            return None

        # 2. Caminhos absolutos são fundamentais para automação COM
        caminho_word = os.path.abspath(caminho_word)
        caminho_pdf = caminho_word.replace(".docx", ".pdf")

        # 3. Conversão direta via Word Application (sem docx2pdf)
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        try:
            doc = word.Documents.Open(caminho_word)
            # 17 é o código para o formato PDF no Word
            doc.SaveAs(caminho_pdf, FileFormat=17)
            doc.Close()
        finally:
            word.Quit()
        
        if os.path.exists(caminho_word):
            os.remove(caminho_word)

        # 4. Abre no navegador
        abrir_no_navegador(caminho_pdf)
        return caminho_pdf
    except Exception as e:
        print(f"Erro na conversão para PDF: {e}")
        return None

def importar_template_arquivo(caminho_origem):
    """Copia um arquivo docx para a pasta de templates."""
    try:
        garantir_pastas()
        nome_arquivo = os.path.basename(caminho_origem)
        if not nome_arquivo.lower().endswith(".docx"):
            return False, "O arquivo deve ser .docx"
        
        destino = os.path.join(TEMPLATES_DIR, nome_arquivo)
        shutil.copy2(caminho_origem, destino)
        return True, f"Template '{nome_arquivo}' importado com sucesso."
    except Exception as e:
        return False, str(e)

# 6. FUNÇÕES PARA EXCEL
def get_templates_excel():
    """Retorna os templates Excel disponíveis para a lista de seleção da interface."""
    garantir_pastas()
    templates = {}
    if not os.path.exists(TEMPLATES_DIR):
        return templates
    for arquivo in sorted(os.listdir(TEMPLATES_DIR)):
        if arquivo.lower().endswith(".xlsx"):
            nome_limpo = os.path.splitext(arquivo)[0]
            templates[nome_limpo] = {
                "nome": nome_limpo.replace("_", " ").title(),
                "xlsx_path": os.path.join(TEMPLATES_DIR, arquivo),
                "template": nome_limpo
            }
    return templates

def preenchercel_excel(cliente_info, tipo_documento, saida_dir=OUTPUT_DIR):
    """Preenche um template Excel com os dados do cliente."""
    try:
        garantir_pastas()
        contexto = criar_contexto_cliente(cliente_info)
        template_path = os.path.join(TEMPLATES_DIR, f"{tipo_documento}.xlsx")
        
        if not os.path.exists(template_path):
            return None, f"Template Excel '{tipo_documento}' não encontrado"

        # Carrega a planilha
        wb = load_workbook(template_path)
        ws = wb.active

        # Percorre todas as células da planilha e substitui {{chave}} pelos valores
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Substitui placeholders no formato {{chave}}
                    valor_original = cell.value
                    for chave, valor in contexto.items():
                        placeholder = f"{{{{{chave}}}}}"
                        if placeholder in valor_original:
                            cell.value = valor_original.replace(placeholder, str(valor))
                            valor_original = cell.value

        # Sanitiza o nome do arquivo
        nome_limpo = "".join([c for c in str(contexto['nome']) if c.isalnum() or c in (' ', '_')]).strip()
        nome_curto = nome_limpo[:50].replace(' ', '_')
        nome_arq = f"{nome_curto}_{datetime.now().strftime('%H%M%S')}.xlsx"
        caminho = os.path.join(saida_dir, nome_arq)
        
        wb.save(caminho)
        return caminho, "Excel gerado com sucesso"
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        return None, f"Erro: {str(e)}"

def gerar_documento_excel(cliente_info, tipo_documento):
    """Gera um documento Excel e o abre automaticamente."""
    try:
        caminho, mensagem = preenchercel_excel(cliente_info, tipo_documento)
        if caminho:
            abrir_no_navegador(caminho)
        return caminho, mensagem
    except Exception as e:
        return None, f"Erro: {str(e)}"

def importar_template_excel(caminho_origem):
    """Copia um arquivo xlsx para a pasta de templates."""
    try:
        garantir_pastas()
        nome_arquivo = os.path.basename(caminho_origem)
        if not nome_arquivo.lower().endswith(".xlsx"):
            return False, "O arquivo deve ser .xlsx"
        
        destino = os.path.join(TEMPLATES_DIR, nome_arquivo)
        shutil.copy2(caminho_origem, destino)
        return True, f"Template Excel '{nome_arquivo}' importado com sucesso."
    except Exception as e:
        return False, str(e)
